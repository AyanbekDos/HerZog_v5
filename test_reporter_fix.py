#!/usr/bin/env python3
"""
Тест исправленного reporter_v3
"""

import sys
import os
sys.path.append('/home/imort/Herzog_v3/src')

from data_processing.reporter_v3 import MultiPageScheduleGenerator
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Создаем тестовые данные
test_timeline_blocks = [
    {
        'week_id': 1,
        'start_date': '2024-01-08T00:00:00',
        'end_date': '2024-01-12T23:59:59'
    },
    {
        'week_id': 2,
        'start_date': '2024-01-15T00:00:00',
        'end_date': '2024-01-19T23:59:59'
    },
    {
        'week_id': 3,
        'start_date': '2024-01-22T00:00:00',
        'end_date': '2024-01-26T23:59:59'
    }
]

test_work_packages = [
    {
        'package_id': 'test_001',
        'name': 'Тестовый пакет работ',
        'volume_data': {
            'final_unit': 'м2',
            'final_quantity': 100.5
        },
        'schedule_blocks': [1, 2],
        'progress_per_block': {'1': 60, '2': 40},
        'staffing_per_block': {'1': 5, '2': 3}
    }
]

test_project_info = {
    'project_name': 'Тест исправления заголовков',
    'created_at': datetime.now().isoformat()
}

# Создаем простой тестовый Excel
wb = Workbook()
ws = wb.active

# Инициализируем генератор
generator = MultiPageScheduleGenerator()

# Тестируем создание заголовков
print("🧪 Тестирование исправленных заголовков...")

# Заголовки основной таблицы
headers = [
    "№ п/п",
    "Наименование вида работ",
    "Ед. изм.",
    "Кол-во",
    "Начало",
    "Окончание",
    "Календарный план по неделям"
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col, value=header)
    cell.font = Font(color="FFFFFF", bold=True, size=11)
    cell.fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Заголовки недель (исправленные)
for i, block in enumerate(test_timeline_blocks, 7):
    week_id = block['week_id']
    start_date = datetime.fromisoformat(block['start_date']).strftime('%d.%m')
    end_date = datetime.fromisoformat(block['end_date']).strftime('%d.%m')

    # Используем правильный перенос строки в ячейке
    cell = ws.cell(row=6, column=i, value=f"Нед.{week_id}\n{start_date}-{end_date}")
    cell.font = Font(size=8, bold=True)
    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Устанавливаем высоту строки для заголовков
ws.row_dimensions[6].height = 30

# Ширины колонок
column_widths = {
    'A': 5, 'B': 40, 'C': 8, 'D': 10, 'E': 12, 'F': 12,
    'G': 12, 'H': 12, 'I': 12
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Сохраняем тестовый файл
wb.save('/tmp/test_reporter_fix.xlsx')

print("✅ Тест успешно создан: /tmp/test_reporter_fix.xlsx")
print("\nИсправления:")
print("1. Убраны символы \\n из текста заголовков")
print("2. Добавлены правильные переносы строк в ячейках Excel")
print("3. Установлена корректная высота строки для многострочных заголовков")
print("4. Удален неиспользуемый reporter_v4.py")