#!/usr/bin/env python3
"""
Тест создания заголовков Excel для проверки символов \n
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Создаем тестовую книгу
wb = Workbook()
ws = wb.active
ws.title = "Тест заголовков"

# Тестируем различные варианты заголовков
test_headers = [
    # Правильный вариант (как в коде)
    "№\nп/п    Наименование\nвида работ    Ед.\nизм.    Кол-во    Начало    Окончание    Календарный план по неделям            ",

    # Правильные заголовки (как должно быть)
    ["№ п/п", "Наименование вида работ", "Ед. изм.", "Кол-во", "Начало", "Окончание", "Календарный план по неделям"]
]

# Строка 1: проблемный заголовок с \n
ws['A1'] = test_headers[0]
ws.merge_cells('A1:G1')

# Строка 3: правильные заголовки по отдельности
for col, header in enumerate(test_headers[1], 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Тестовые заголовки недель
timeline_data = [
    {"week_id": 1, "start": "08.01", "end": "12.01"},
    {"week_id": 2, "start": "15.01", "end": "19.01"},
    {"week_id": 3, "start": "22.01", "end": "26.01"},
    {"week_id": 4, "start": "29.01", "end": "30.01"}
]

# Строка 4: заголовки недель (правильный способ)
for i, week in enumerate(timeline_data, 7):  # Начинаем с колонки G
    week_header = f"Нед.{week['week_id']}\n{week['start']}-{week['end']}"
    cell = ws.cell(row=4, column=i, value=week_header)
    cell.font = Font(size=8, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Ширина колонок
column_widths = {'A': 5, 'B': 40, 'C': 8, 'D': 10, 'E': 12, 'F': 12}
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Ширина колонок с неделями
for i in range(7, 11):  # Колонки G, H, I, J
    from openpyxl.utils import get_column_letter
    col_letter = get_column_letter(i)
    ws.column_dimensions[col_letter].width = 10

# Сохраняем тестовый файл
wb.save('/tmp/test_headers.xlsx')
print("✅ Тестовый файл создан: /tmp/test_headers.xlsx")
print("\nСравните:")
print(f"1. Проблемный заголовок: '{test_headers[0][:50]}...'")
print(f"2. Правильные заголовки: {test_headers[1]}")