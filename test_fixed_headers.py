#!/usr/bin/env python3
"""
Тест исправленного формирования заголовков Excel
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# Создаем тестовую книгу
wb = Workbook()
ws = wb.active
ws.title = "Исправленные заголовки"

# Заголовки основной таблицы (правильный способ)
headers = [
    "№ п/п",
    "Наименование вида работ",
    "Ед. изм.",
    "Кол-во",
    "Начало",
    "Окончание",
    "Календарный план по неделям"
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=5, column=col, value=header)
    cell.font = Font(color="FFFFFF", bold=True, size=11)
    cell.fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Исправленные заголовки недель (с переносом строки внутри ячейки)
timeline_data = [
    {"week_id": 1, "start": "08.01", "end": "12.01"},
    {"week_id": 2, "start": "15.01", "end": "19.01"},
    {"week_id": 3, "start": "22.01", "end": "26.01"},
    {"week_id": 4, "start": "29.01", "end": "30.01"}
]

for i, week in enumerate(timeline_data, 7):  # Начинаем с колонки G
    # Используем символ \n внутри ячейки для переноса строки
    week_header = f"Нед.{week['week_id']}\n{week['start']}-{week['end']}"
    cell = ws.cell(row=6, column=i, value=week_header)
    cell.font = Font(size=8, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

# Ширина колонок
column_widths = {
    'A': 5, 'B': 40, 'C': 8, 'D': 10, 'E': 12, 'F': 12,
    'G': 10, 'H': 10, 'I': 10, 'J': 10
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Устанавливаем высоту строки для правильного отображения многострочных ячеек
ws.row_dimensions[6].height = 30

# Сохраняем исправленный тестовый файл
wb.save('/tmp/test_fixed_headers.xlsx')
print("✅ Исправленный тестовый файл создан: /tmp/test_fixed_headers.xlsx")
print("\nТеперь заголовки недель должны корректно отображаться с переносами строк внутри ячеек")
print("а не в виде плоского текста с символами \\n")