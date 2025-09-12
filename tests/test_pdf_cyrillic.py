#!/usr/bin/env python3
"""
Тест исправления кириллицы в PDF экспорте
"""

import sys
import os

# Добавляем путь к модулям Herzog
sys.path.append('/home/imort/Herzog_v3')

from src.data_processing.pdf_exporter import PDFExporter

def test_pdf_cyrillic():
    """Тестируем PDF с кириллицей"""
    
    print("🧪 Тестирование PDF экспорта с кириллицей...")
    
    # Создаем простой тестовый Excel файл с кириллицей
    import openpyxl
    from datetime import datetime
    
    # Создаем временный Excel файл
    test_excel_path = '/tmp/test_cyrillic.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Календарный график"
    
    # Заполняем данными с кириллицей
    ws['A1'] = "№"
    ws['B1'] = "Наименование работ"
    ws['C1'] = "Единица измерения" 
    ws['D1'] = "Объем"
    
    ws['A2'] = 1
    ws['B2'] = "Демонтаж перегородок"
    ws['C2'] = "м²"
    ws['D2'] = 150.5
    
    ws['A3'] = 2
    ws['B3'] = "Монтаж стяжки пола"
    ws['C3'] = "м²"  
    ws['D3'] = 85.2
    
    ws['A4'] = 3
    ws['B4'] = "Устройство чистового покрытия"
    ws['C4'] = "м²"
    ws['D4'] = 85.2
    
    # Сохраняем Excel
    wb.save(test_excel_path)
    print(f"📄 Создан тестовый Excel: {test_excel_path}")
    
    # Тестируем PDF экспорт
    try:
        exporter = PDFExporter()
        pdf_path = exporter.export_excel_to_pdf(test_excel_path, '/tmp', 'pdf')
        
        if os.path.exists(pdf_path):
            file_size = os.path.getsize(pdf_path)
            print(f"✅ PDF создан успешно: {pdf_path}")
            print(f"📊 Размер файла: {file_size} байт")
            
            # Проверяем содержимое PDF (простая проверка)
            with open(pdf_path, 'rb') as f:
                content = f.read()
                if b'PDF' in content[:10]:
                    print("✅ Файл является корректным PDF")
                else:
                    print("⚠️ Файл может быть поврежден")
                    
        else:
            print("❌ PDF файл не был создан")
            
    except Exception as e:
        print(f"❌ Ошибка создания PDF: {e}")
        import traceback
        traceback.print_exc()
    
    # Очищаем временные файлы
    try:
        if os.path.exists(test_excel_path):
            os.remove(test_excel_path)
    except:
        pass

if __name__ == "__main__":
    test_pdf_cyrillic()