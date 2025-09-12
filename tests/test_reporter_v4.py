"""
Тестовый файл для проверки reporter_v4.py
с улучшенным UI/UX и scheduling_reasoning данными
"""

import os
import sys
import logging

# Добавляем путь к модулю
sys.path.append('/home/imort/Herzog_v3/src/data_processing')

from reporter_v4 import generate_professional_excel_report

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

def test_professional_reporter():
    """
    Тестирование нового профессионального reporter_v4
    """
    print("🧪 ТЕСТИРОВАНИЕ ПРОФЕССИОНАЛЬНОГО REPORTER V4")
    print("=" * 60)
    
    # Тестовые пути
    test_cases = [
        "/home/imort/Herzog_v3/projects/test/b4338a45/true.json",  # С scheduling_reasoning
        "/home/imort/Herzog_v3/projects/test/d3f0a7a1/true.json",  # Альтернативный
        "/home/imort/Herzog_v3/projects/test/9d778a7f/true.json"   # Резервный
    ]
    
    output_dir = "/tmp"
    
    for i, test_input in enumerate(test_cases, 1):
        print(f"\n📊 ТЕСТ {i}: {os.path.basename(os.path.dirname(test_input))}")
        print("-" * 40)
        
        if not os.path.exists(test_input):
            print(f"❌ Тестовый файл не найден: {test_input}")
            continue
        
        # Проверяем наличие scheduler данных
        scheduler_path = os.path.join(os.path.dirname(test_input), '7_scheduler_and_staffer', 'llm_response.json')
        has_scheduler_data = os.path.exists(scheduler_path)
        print(f"📅 Данные планирования: {'✅ Найдены' if has_scheduler_data else '❌ Отсутствуют'}")
        
        if has_scheduler_data:
            # Проверяем структуру scheduler данных
            try:
                import json
                with open(scheduler_path, 'r', encoding='utf-8') as f:
                    scheduler_data = json.load(f)
                    
                success = scheduler_data.get('success', False)
                packages = scheduler_data.get('response', {}).get('scheduled_packages', [])
                
                print(f"📊 Статус планирования: {'✅ Успешно' if success else '❌ Ошибка'}")
                print(f"📦 Пакетов с обоснованиями: {len(packages)}")
                
                # Проверяем наличие reasoning
                reasoning_count = 0
                for pkg in packages:
                    if pkg.get('scheduling_reasoning'):
                        reasoning_count += 1
                
                print(f"🧠 Пакетов с reasoning: {reasoning_count}")
                
            except Exception as e:
                print(f"❌ Ошибка чтения данных планирования: {e}")
        
        # Запускаем генерацию отчета
        try:
            print(f"🚀 Генерация профессионального отчета...")
            result_file = generate_professional_excel_report(test_input, output_dir)
            
            if os.path.exists(result_file):
                file_size = os.path.getsize(result_file) // 1024  # KB
                print(f"✅ УСПЕХ! Отчет создан: {result_file}")
                print(f"📊 Размер файла: {file_size} KB")
                
                # Проверяем структуру файла
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(result_file)
                    sheet_names = wb.sheetnames
                    
                    print(f"📋 Листы в файле ({len(sheet_names)}):")
                    for sheet in sheet_names:
                        ws = wb[sheet]
                        rows = ws.max_row
                        cols = ws.max_column
                        print(f"  • {sheet}: {rows} строк, {cols} колонок")
                        
                        # Специальная проверка для листа "Планирование и Обоснования"
                        if "Планирование" in sheet and has_scheduler_data:
                            print(f"    🎯 Новый лист с обоснованиями обнаружен!")
                    
                    wb.close()
                    
                except Exception as e:
                    print(f"⚠️ Ошибка анализа структуры файла: {e}")
                
                print(f"🎉 ТЕСТ {i} ПРОЙДЕН!")
                
            else:
                print(f"❌ ТЕСТ {i} ПРОВАЛЕН: Файл не создан")
                
        except Exception as e:
            print(f"❌ ТЕСТ {i} ПРОВАЛЕН: {e}")
            import traceback
            traceback.print_exc()
    
    print("\n" + "=" * 60)
    print("📊 ТЕСТИРОВАНИЕ ЗАВЕРШЕНО")

def analyze_test_data():
    """
    Анализирует доступные тестовые данные
    """
    print("\n🔍 АНАЛИЗ ДОСТУПНЫХ ТЕСТОВЫХ ДАННЫХ")
    print("=" * 60)
    
    projects_dir = "/home/imort/Herzog_v3/projects"
    
    if not os.path.exists(projects_dir):
        print("❌ Папка projects не найдена")
        return
    
    # Найдем все true.json файлы
    true_json_files = []
    for root, dirs, files in os.walk(projects_dir):
        if 'true.json' in files:
            true_json_files.append(os.path.join(root, 'true.json'))
    
    print(f"📊 Найдено файлов true.json: {len(true_json_files)}")
    
    # Анализируем каждый файл
    for i, filepath in enumerate(true_json_files[:5], 1):  # Ограничим до 5 для краткости
        project_id = os.path.basename(os.path.dirname(filepath))
        print(f"\n📁 ПРОЕКТ {i}: {project_id}")
        print("-" * 30)
        
        try:
            import json
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Базовая информация
            meta = data.get('meta', {})
            results = data.get('results', {})
            work_packages = results.get('work_packages', [])
            timeline_blocks = data.get('timeline_blocks', [])
            
            print(f"📊 Пакетов работ: {len(work_packages)}")
            print(f"⏰ Временных блоков: {len(timeline_blocks)}")
            print(f"📅 Версия структуры: {meta.get('structure_version', '1.0')}")
            
            # Проверяем scheduler данные
            scheduler_path = os.path.join(os.path.dirname(filepath), '7_scheduler_and_staffer', 'llm_response.json')
            if os.path.exists(scheduler_path):
                with open(scheduler_path, 'r', encoding='utf-8') as f:
                    scheduler_data = json.load(f)
                
                success = scheduler_data.get('success', False)
                scheduled_packages = scheduler_data.get('response', {}).get('scheduled_packages', [])
                
                reasoning_count = sum(1 for pkg in scheduled_packages if pkg.get('scheduling_reasoning'))
                
                print(f"🧠 Планирование: {'✅ Да' if success else '❌ Нет'}")
                print(f"💭 С обоснованиями: {reasoning_count}/{len(scheduled_packages)} пакетов")
                
                if reasoning_count > 0:
                    print(f"🎯 ПОДХОДИТ ДЛЯ ТЕСТИРОВАНИЯ!")
            else:
                print("🧠 Планирование: ❌ Данные отсутствуют")
            
        except Exception as e:
            print(f"❌ Ошибка анализа: {e}")

if __name__ == "__main__":
    print("🚀 ЗАПУСК ТЕСТИРОВАНИЯ REPORTER V4")
    print("Профессиональный Excel отчет с современным UI/UX")
    print("Включает scheduling_reasoning данные из AI-агента")
    
    # Сначала анализируем доступные данные
    analyze_test_data()
    
    # Затем запускаем тестирование
    test_professional_reporter()
    
    print("\n✨ ТЕСТИРОВАНИЕ ЗАВЕРШЕНО! Проверьте созданные файлы в /tmp")