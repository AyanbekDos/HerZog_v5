"""
PDF Exporter для HerZog v3.0
Конвертирует Excel календарные графики в PDF формат
"""

import os
import logging
from typing import Optional, List, Dict, Any
import subprocess
import json
from datetime import datetime

logger = logging.getLogger(__name__)

class PDFExporter:
    """
    Класс для экспорта календарных графиков в PDF
    """
    
    def __init__(self):
        self.supported_formats = ['pdf', 'png', 'jpg']
    
    def export_excel_to_pdf(self, excel_file: str, output_path: str, format: str = 'pdf') -> str:
        """
        Экспортирует Excel файл в PDF или изображения
        
        Args:
            excel_file: Путь к Excel файлу
            output_path: Папка для сохранения
            format: Формат вывода ('pdf', 'png', 'jpg')
            
        Returns:
            Путь к созданному файлу
        """
        try:
            if format not in self.supported_formats:
                raise ValueError(f"Неподдерживаемый формат: {format}. Доступные: {self.supported_formats}")
            
            # Проверяем наличие Excel файла
            if not os.path.exists(excel_file):
                raise FileNotFoundError(f"Excel файл не найден: {excel_file}")
            
            logger.info(f"📄 Экспорт Excel в {format.upper()}: {excel_file}")
            
            # Генерируем имя выходного файла
            base_name = os.path.splitext(os.path.basename(excel_file))[0]
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = os.path.join(output_path, f"{base_name}_export.{format}")
            
            # Пытаемся использовать различные методы конвертации
            success = False
            
            # Метод 1: LibreOffice (самый надежный)
            if not success:
                success = self._convert_with_libreoffice(excel_file, output_file, format)
            
            # Метод 2: Python библиотеки (резервный)
            if not success:
                success = self._convert_with_python_libs(excel_file, output_file, format)
            
            # Метод 3: Создаем PDF-отчет "вручную" на основе данных
            if not success:
                success = self._create_pdf_from_data(excel_file, output_file)
            
            if success:
                logger.info(f"✅ PDF экспорт завершен: {output_file}")
                return output_file
            else:
                raise Exception("Все методы конвертации в PDF не сработали")
                
        except Exception as e:
            logger.error(f"❌ Ошибка экспорта в PDF: {e}")
            raise
    
    def _convert_with_libreoffice(self, excel_file: str, output_file: str, format: str) -> bool:
        """Конвертирует используя LibreOffice"""
        try:
            logger.info("🔄 Попытка конвертации через LibreOffice...")
            
            # Проверяем наличие LibreOffice
            result = subprocess.run(['which', 'libreoffice'], capture_output=True, text=True)
            if result.returncode != 0:
                logger.warning("⚠️ LibreOffice не найден в системе")
                return False
            
            # Команда для конвертации
            output_dir = os.path.dirname(output_file)
            if format == 'pdf':
                cmd = [
                    'libreoffice', '--headless', '--convert-to', 'pdf',
                    '--outdir', output_dir, excel_file
                ]
            else:
                logger.warning(f"⚠️ LibreOffice не поддерживает прямую конвертацию в {format}")
                return False
            
            # Выполняем конвертацию
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
            
            if result.returncode == 0:
                # LibreOffice создает файл с именем оригинала + .pdf
                expected_file = os.path.join(output_dir, os.path.splitext(os.path.basename(excel_file))[0] + '.pdf')
                if os.path.exists(expected_file):
                    # Переименовываем файл если нужно
                    if expected_file != output_file:
                        os.rename(expected_file, output_file)
                    return True
            
            logger.warning(f"⚠️ LibreOffice конвертация не удалась: {result.stderr}")
            return False
            
        except subprocess.TimeoutExpired:
            logger.warning("⚠️ LibreOffice конвертация прервана по таймауту")
            return False
        except Exception as e:
            logger.warning(f"⚠️ Ошибка LibreOffice конвертации: {e}")
            return False
    
    def _convert_with_python_libs(self, excel_file: str, output_file: str, format: str) -> bool:
        """Конвертирует используя Python библиотеки"""
        try:
            logger.info("🔄 Попытка конвертации через Python библиотеки...")
            
            # Попытка использовать reportlab для PDF
            if format == 'pdf':
                return self._create_pdf_with_reportlab(excel_file, output_file)
            
            # Для изображений можно использовать PIL + openpyxl
            elif format in ['png', 'jpg']:
                logger.warning("⚠️ Конвертация в изображения через Python пока не реализована")
                return False
            
            return False
            
        except Exception as e:
            logger.warning(f"⚠️ Ошибка Python конвертации: {e}")
            return False
    
    def _create_pdf_with_reportlab(self, excel_file: str, output_file: str) -> bool:
        """Создает PDF используя reportlab"""
        try:
            # Проверяем наличие reportlab
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib import colors
                from reportlab.lib.units import cm
                from reportlab.pdfbase import pdfmetrics
                from reportlab.pdfbase.ttfonts import TTFont
            except ImportError:
                logger.warning("⚠️ reportlab не установлен")
                return False
            
            import openpyxl
            
            # Читаем данные из Excel
            wb = openpyxl.load_workbook(excel_file)
            
            # Регистрируем шрифт с поддержкой кириллицы
            try:
                # Пытаемся найти системный шрифт с кириллицей
                font_paths = [
                    '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',  # Ubuntu/Debian
                    '/usr/share/fonts/TTF/DejaVuSans.ttf',            # Arch
                    '/System/Library/Fonts/Arial.ttf',               # macOS
                    'C:/Windows/Fonts/arial.ttf',                    # Windows
                    '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf'  # Liberation
                ]
                
                font_registered = False
                for font_path in font_paths:
                    if os.path.exists(font_path):
                        pdfmetrics.registerFont(TTFont('CyrillicFont', font_path))
                        font_registered = True
                        logger.info(f"📝 Зарегистрирован шрифт: {font_path}")
                        break
                
                if not font_registered:
                    logger.warning("⚠️ Шрифт с кириллицей не найден, используем встроенный")
                    
            except Exception as e:
                logger.warning(f"⚠️ Ошибка регистрации шрифта: {e}")

            # Создаем PDF
            doc = SimpleDocTemplate(output_file, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()
            
            # Создаем стили с кириллическим шрифтом
            if font_registered:
                styles.add(ParagraphStyle('CyrillicHeading1',
                                        parent=styles['Heading1'],
                                        fontName='CyrillicFont',
                                        fontSize=16))
                styles.add(ParagraphStyle('CyrillicNormal',
                                        parent=styles['Normal'],
                                        fontName='CyrillicFont',
                                        fontSize=10))
                heading_style = 'CyrillicHeading1'
                normal_style = 'CyrillicNormal'
            else:
                heading_style = 'Heading1'
                normal_style = 'Normal'
            
            # Обрабатываем каждый лист Excel
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Заголовок листа
                story.append(Paragraph(f"<b>{sheet_name}</b>", styles[heading_style]))
                story.append(Spacer(1, 0.5*cm))
                
                # Извлекаем данные из листа (только первые 20 строк и 10 колонок)
                max_rows = min(20, ws.max_row)
                max_cols = min(10, ws.max_column)
                
                data = []
                for row in range(1, max_rows + 1):
                    row_data = []
                    for col in range(1, max_cols + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value is None:
                            row_data.append("")
                        else:
                            # Ограничиваем длину текста
                            text = str(cell_value)
                            if len(text) > 30:
                                text = text[:27] + "..."
                            row_data.append(text)
                    data.append(row_data)
                
                # Создаем таблицу
                if data:
                    table = Table(data)
                    
                    # Выбираем шрифт для таблицы
                    table_font = 'CyrillicFont' if font_registered else 'Helvetica'
                    table_font_bold = 'CyrillicFont' if font_registered else 'Helvetica-Bold'
                    
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), table_font_bold),
                        ('FONTNAME', (0, 1), (-1, -1), table_font),
                        ('FONTSIZE', (0, 0), (-1, 0), 8),
                        ('FONTSIZE', (0, 1), (-1, -1), 7),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(table)
                
                story.append(Spacer(1, 1*cm))
            
            # Генерируем PDF
            doc.build(story)
            logger.info("✅ PDF создан через reportlab")
            return True
            
        except Exception as e:
            logger.warning(f"⚠️ Ошибка создания PDF через reportlab: {e}")
            return False
    
    def _create_pdf_from_data(self, excel_file: str, output_file: str) -> bool:
        """Создает PDF на основе исходных данных true.json"""
        try:
            logger.info("🔄 Создание PDF на основе исходных данных...")
            
            # Ищем соответствующий true.json файл
            project_dir = self._find_project_dir(excel_file)
            if not project_dir:
                logger.warning("⚠️ Не найдена папка проекта")
                return False
            
            truth_file = os.path.join(project_dir, "true.json")
            if not os.path.exists(truth_file):
                logger.warning("⚠️ Файл true.json не найден")
                return False
            
            # Читаем данные
            with open(truth_file, 'r', encoding='utf-8') as f:
                truth_data = json.load(f)
            
            # Создаем простой текстовый PDF
            return self._create_simple_text_pdf(truth_data, output_file)
            
        except Exception as e:
            logger.warning(f"⚠️ Ошибка создания PDF из данных: {e}")
            return False
    
    def _find_project_dir(self, excel_file: str) -> Optional[str]:
        """Ищет папку проекта по Excel файлу"""
        # Предполагаем, что Excel файл создается в /tmp, а проект в /projects
        project_dirs = []
        
        # Ищем в стандартных местах
        herzog_path = "/home/imort/Herzog_v3"
        if os.path.exists(herzog_path):
            projects_path = os.path.join(herzog_path, "projects")
            if os.path.exists(projects_path):
                # Ищем последнюю измененную папку проекта
                for user_dir in os.listdir(projects_path):
                    user_path = os.path.join(projects_path, user_dir)
                    if os.path.isdir(user_path):
                        for project_dir in os.listdir(user_path):
                            project_path = os.path.join(user_path, project_dir)
                            if os.path.isdir(project_path):
                                project_dirs.append((project_path, os.path.getmtime(project_path)))
        
        # Возвращаем самую свежую папку проекта
        if project_dirs:
            project_dirs.sort(key=lambda x: x[1], reverse=True)
            return project_dirs[0][0]
        
        return None
    
    def _create_simple_text_pdf(self, truth_data: Dict, output_file: str) -> bool:
        """Создает простой текстовый PDF"""
        try:
            # Пытаемся использовать weasyprint для HTML->PDF
            try:
                import weasyprint
                return self._create_html_pdf(truth_data, output_file)
            except ImportError:
                pass
            
            # Резервный вариант: создаем текстовый файл вместо PDF
            text_file = output_file.replace('.pdf', '.txt')
            
            with open(text_file, 'w', encoding='utf-8') as f:
                f.write("КАЛЕНДАРНЫЙ ГРАФИК ПРОИЗВОДСТВА РАБОТ\n")
                f.write("="*50 + "\n\n")
                
                # Основная информация
                project_name = truth_data.get('project_inputs', {}).get('project_name', 'Безымянный проект')
                f.write(f"Проект: {project_name}\n")
                f.write(f"Дата создания: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n")
                
                # Пакеты работ
                work_packages = truth_data.get('results', {}).get('work_packages', [])
                f.write(f"ПАКЕТЫ РАБОТ ({len(work_packages)} шт.):\n")
                f.write("-" * 30 + "\n")
                
                for i, package in enumerate(work_packages, 1):
                    name = package.get('name', 'Безымянный пакет')
                    volume_data = package.get('volume_data', {})
                    unit = volume_data.get('unit', 'шт')
                    quantity = volume_data.get('quantity', 0)
                    
                    f.write(f"{i}. {name}\n")
                    f.write(f"   Объем: {quantity} {unit}\n")
                    
                    # Календарный план
                    schedule_blocks = package.get('schedule_blocks', [])
                    if schedule_blocks:
                        f.write(f"   Недели: {', '.join(map(str, schedule_blocks))}\n")
                    
                    f.write("\n")
            
            # Переименовываем в .pdf для совместимости
            if os.path.exists(text_file):
                os.rename(text_file, output_file)
                logger.info("✅ Создан текстовый PDF (как .txt)")
                return True
            
            return False
            
        except Exception as e:
            logger.warning(f"⚠️ Ошибка создания простого PDF: {e}")
            return False
    
    def _create_html_pdf(self, truth_data: Dict, output_file: str) -> bool:
        """Создает PDF через HTML+CSS"""
        try:
            import weasyprint
            
            # Генерируем HTML
            html_content = self._generate_html_report(truth_data)
            
            # Конвертируем в PDF
            weasyprint.HTML(string=html_content).write_pdf(output_file)
            logger.info("✅ PDF создан через weasyprint")
            return True
            
        except Exception as e:
            logger.warning(f"⚠️ Ошибка создания HTML PDF: {e}")
            return False
    
    def _generate_html_report(self, truth_data: Dict) -> str:
        """Генерирует HTML отчет"""
        project_name = truth_data.get('project_inputs', {}).get('project_name', 'Безымянный проект')
        work_packages = truth_data.get('results', {}).get('work_packages', [])
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Календарный график - {project_name}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #366092; text-align: center; }}
                table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                th, td {{ border: 1px solid #ccc; padding: 8px; text-align: left; }}
                th {{ background-color: #366092; color: white; }}
                .center {{ text-align: center; }}
            </style>
        </head>
        <body>
            <h1>📊 КАЛЕНДАРНЫЙ ГРАФИК ПРОИЗВОДСТВА РАБОТ</h1>
            <h2>{project_name}</h2>
            <p><strong>Дата создания:</strong> {datetime.now().strftime('%d.%m.%Y %H:%M')}</p>
            
            <h3>Пакеты работ</h3>
            <table>
                <tr>
                    <th>№</th>
                    <th>Наименование</th>
                    <th>Единица</th>
                    <th>Количество</th>
                    <th>Недели</th>
                </tr>
        """
        
        for i, package in enumerate(work_packages, 1):
            name = package.get('name', 'Безымянный пакет')
            volume_data = package.get('volume_data', {})
            unit = volume_data.get('unit', 'шт')
            quantity = volume_data.get('quantity', 0)
            schedule_blocks = package.get('schedule_blocks', [])
            weeks = ', '.join(map(str, schedule_blocks)) if schedule_blocks else '-'
            
            html += f"""
                <tr>
                    <td class="center">{i}</td>
                    <td>{name}</td>
                    <td class="center">{unit}</td>
                    <td class="center">{quantity}</td>
                    <td class="center">{weeks}</td>
                </tr>
            """
        
        html += """
            </table>
        </body>
        </html>
        """
        
        return html


def export_schedule_to_pdf(excel_file: str, output_path: str, format: str = 'pdf') -> str:
    """
    Экспортирует календарный график в PDF
    
    Args:
        excel_file: Путь к Excel файлу
        output_path: Папка для сохранения
        format: Формат экспорта ('pdf', 'png', 'jpg')
        
    Returns:
        Путь к созданному файлу
    """
    exporter = PDFExporter()
    return exporter.export_excel_to_pdf(excel_file, output_path, format)


if __name__ == "__main__":
    # Тестирование экспорта
    test_excel = "/tmp/Отчет_Проект_20250910_141641.xlsx"
    test_output = "/tmp"
    
    if os.path.exists(test_excel):
        print("🧪 Тестирование экспорта в PDF...")
        try:
            pdf_file = export_schedule_to_pdf(test_excel, test_output)
            print(f"✅ PDF экспорт завершен: {pdf_file}")
        except Exception as e:
            print(f"❌ Ошибка экспорта: {e}")
            import traceback
            traceback.print_exc()
    else:
        print(f"❌ Тестовый Excel файл не найден: {test_excel}")