"""
НОВЫЙ Модуль REPORTER v3 для HerZog v3.0
Создание МНОГОСТРАНИЧНОГО Excel отчета с листами:
- 📊 График (календарный план Gantt)
- 📋 Пакеты работ (детальная информация)
- 🧮 Логика расчетов (техническая информация)
"""

import json
import logging
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, NamedStyle
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Any

logger = logging.getLogger(__name__)

class MultiPageScheduleGenerator:
    """
    Генератор многостраничного календарного графика в стиле HerZog v3.0
    """
    
    def __init__(self):
        # Современные стили для Excel
        self.header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")  # Темно-синий
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.subheader_fill = PatternFill(start_color="4A628A", end_color="4A628A", fill_type="solid")  # Средний синий
        self.subheader_font = Font(color="FFFFFF", bold=True, size=10)
        
        # Прогресс-бары с градиентами
        self.progress_high = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")  # Зеленый
        self.progress_medium = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")  # Оранжевый
        self.progress_low = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")  # Красный
        
        # Акцентные цвета
        self.info_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Светло-голубой
        self.logic_fill = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")  # Светло-зеленый
        self.reasoning_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")  # Светло-оранжевый
        self.warning_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")  # Светло-красный
        
        # Границы
        self.border = Border(
            left=Side(style='thin', color='CCCCCC'), 
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'), 
            bottom=Side(style='thin', color='CCCCCC')
        )
        self.thick_border = Border(
            left=Side(style='medium', color='2E4057'), 
            right=Side(style='medium', color='2E4057'),
            top=Side(style='medium', color='2E4057'), 
            bottom=Side(style='medium', color='2E4057')
        )
        
        # Выравнивание
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.right_align = Alignment(horizontal='right', vertical='center')
        self.top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    def generate_multipage_excel(self, input_file: str, output_path: str) -> str:
        """
        Генерация многостраничного Excel отчета из true.json
        
        Args:
            input_file: Путь к файлу true.json
            output_path: Папка для сохранения
            
        Returns:
            Путь к созданному файлу
        """
        try:
            # Читаем данные из true.json
            with open(input_file, 'r', encoding='utf-8') as f:
                truth_data = json.load(f)
            
            # Определяем версию структуры
            structure_version = truth_data.get('meta', {}).get('structure_version', '1.0')
            logger.info(f"📊 Обработка true.json версии {structure_version}")
            
            # Извлекаем данные в зависимости от версии
            if structure_version == "2.0":
                extracted_data = self._extract_data_v2(truth_data)
            else:
                extracted_data = self._extract_data_v1(truth_data)
            
            work_packages = extracted_data['work_packages']
            timeline_blocks = extracted_data['timeline_blocks']
            project_info = extracted_data['project_info']
            
            if not work_packages:
                raise Exception("Нет пакетов работ для создания календарного графика")
            
            if not timeline_blocks:
                raise Exception("Нет временных блоков для календарного графика")
            
            logger.info(f"📊 Создание отчета для {len(work_packages)} пакетов на {len(timeline_blocks)} недель")
            
            # Загружаем scheduling_reasoning данные
            scheduling_data = self._load_scheduling_reasoning(input_file)
            
            # Создаем Excel с несколькими листами
            wb = Workbook()
            
            # Лист 1: 📊 График (Gantt)
            ws_schedule = wb.active
            ws_schedule.title = "📊 График"
            self._create_schedule_sheet(ws_schedule, work_packages, timeline_blocks, project_info, scheduling_data)
            
            # Лист 2: 📅 Планирование и Обоснования (НОВЫЙ!)
            ws_reasoning = wb.create_sheet("📅 Планирование")
            self._create_reasoning_sheet(ws_reasoning, work_packages, timeline_blocks, project_info, scheduling_data)
            
            # Лист 3: 📋 Пакеты работ
            ws_packages = wb.create_sheet("📋 Пакеты работ")
            self._create_packages_sheet(ws_packages, work_packages, project_info)
            
            # Лист 4: 🧮 Логика расчетов  
            ws_logic = wb.create_sheet("🧮 Логика расчетов")
            self._create_logic_sheet(ws_logic, work_packages, project_info, extracted_data)
            
            # Сохраняем файл
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            project_name = project_info.get('project_name', 'project').replace(' ', '_')
            filename = f"Отчет_{project_name}_{timestamp}.xlsx"
            output_file = os.path.join(output_path, filename)
            
            wb.save(output_file)
            logger.info(f"✅ Многостраничный отчет сохранен: {output_file}")
            
            return output_file
            
        except Exception as e:
            logger.error(f"❌ Ошибка создания многостраничного отчета: {e}")
            import traceback
            logger.error(f"📋 Полная трассировка ошибки:\\n{traceback.format_exc()}")
            raise
    
    def _load_scheduling_reasoning(self, input_file: str) -> Dict[str, Any]:
        """
        Загружает данные scheduling_reasoning из папки scheduler_and_staffer
        """
        try:
            # Определяем путь к папке scheduler_and_staffer
            project_folder = os.path.dirname(input_file)
            scheduler_response_path = os.path.join(project_folder, "7_scheduler_and_staffer", "llm_response.json")
            
            if not os.path.exists(scheduler_response_path):
                logger.warning(f"Файл scheduling_reasoning не найден: {scheduler_response_path}")
                return {}
            
            with open(scheduler_response_path, 'r', encoding='utf-8') as f:
                scheduler_data = json.load(f)
            
            if not scheduler_data.get('success', False):
                logger.warning("Scheduler response не содержит успешного результата")
                return {}
            
            # Парсим JSON из response
            raw_response = scheduler_data.get('raw_text', scheduler_data.get('response', ''))
            if isinstance(raw_response, str):
                try:
                    # Очищаем от markdown если есть
                    if raw_response.strip().startswith('```'):
                        # Убираем ```json в начале и ``` в конце
                        lines = raw_response.strip().split('\n')
                        if lines[0].startswith('```'):
                            lines = lines[1:]
                        if lines[-1].strip() == '```':
                            lines = lines[:-1]
                        raw_response = '\n'.join(lines)
                    
                    parsed_response = json.loads(raw_response)
                except json.JSONDecodeError as e:
                    logger.warning(f"Не удалось распарсить scheduler response JSON: {e}")
                    logger.warning(f"Raw response sample: {raw_response[:200]}...")
                    return {}
            else:
                parsed_response = raw_response
            
            # Извлекаем scheduled_packages с reasoning
            scheduled_packages = parsed_response.get('scheduled_packages', [])
            
            # Создаем словарь для быстрого поиска
            reasoning_dict = {}
            for package in scheduled_packages:
                package_id = package.get('package_id')
                if package_id and 'scheduling_reasoning' in package:
                    reasoning_dict[package_id] = {
                        'scheduling_reasoning': package['scheduling_reasoning'],
                        'schedule_blocks': package.get('schedule_blocks', []),
                        'progress_per_block': package.get('progress_per_block', {}),
                        'staffing_per_block': package.get('staffing_per_block', {})
                    }
            
            logger.info(f"✅ Загружено обоснований планирования для {len(reasoning_dict)} пакетов")
            return reasoning_dict
            
        except Exception as e:
            logger.error(f"❌ Ошибка загрузки scheduling_reasoning: {e}")
            return {}
    
    def _extract_data_v2(self, truth_data: Dict) -> Dict[str, Any]:
        """Извлекает данные из структуры v2.0"""
        return {
            'work_packages': truth_data.get('results', {}).get('work_packages', []),
            'timeline_blocks': truth_data.get('timeline_blocks', []),
            'project_info': {
                'project_name': truth_data.get('meta', {}).get('project_name', 'Проект'),
                'created_at': truth_data.get('meta', {}).get('created_at'),
                'structure_version': '2.0'
            },
            'user_inputs': truth_data.get('user_inputs', {}),
            'pipeline_status': truth_data.get('pipeline', {})
        }
    
    def _extract_data_v1(self, truth_data: Dict) -> Dict[str, Any]:
        """Извлекает данные из структуры v1.0"""
        return {
            'work_packages': truth_data.get('results', {}).get('work_packages', []),
            'timeline_blocks': truth_data.get('timeline_blocks', []),
            'project_info': {
                'project_name': truth_data.get('project_inputs', {}).get('project_name', 'Проект'),
                'created_at': truth_data.get('metadata', {}).get('created_at'),
                'structure_version': '1.0'
            },
            'user_inputs': truth_data.get('project_inputs', {}),
            'pipeline_status': truth_data.get('metadata', {}).get('pipeline_status', [])
        }
    
    def _create_schedule_sheet(self, ws, work_packages: List[Dict], timeline_blocks: List[Dict], project_info: Dict, scheduling_data: Dict = {}):
        """Создает лист с календарным графиком (Gantt)"""
        
        # Заголовок документа
        ws['A1'] = "📊 КАЛЕНДАРНЫЙ ГРАФИК ПРОИЗВОДСТВА РАБОТ"
        ws.merge_cells('A1:G1')
        ws['A1'].font = Font(bold=True, size=16, color="366092")
        ws['A1'].alignment = self.center_align
        
        # Информация о проекте
        ws['A2'] = f"Проект: {project_info.get('project_name', 'Безымянный проект')}"
        ws.merge_cells('A2:G2')
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = self.center_align
        
        ws['A3'] = f"Создан: {self._format_datetime(project_info.get('created_at'))}"
        ws.merge_cells('A3:G3')
        ws['A3'].font = Font(size=10)
        ws['A3'].alignment = self.center_align
        
        # Заголовки колонок (строка 5)
        headers = [
            "№\\nп/п",
            "Наименование\\nвида работ", 
            "Ед.\\nизм.",
            "Кол-во",
            "Начало",
            "Окончание",
            "Календарный план по неделям"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        
        # Объединяем "Календарный план" на нужное количество колонок
        timeline_cols = len(timeline_blocks)
        if timeline_cols > 1:
            ws.merge_cells(f'G5:{get_column_letter(6 + timeline_cols)}5')
        
        # Заголовки недель (строка 6)
        for i, block in enumerate(timeline_blocks, 7):
            week_id = block.get('week_id', block.get('block_id', i-6))
            start_date = datetime.fromisoformat(block['start_date']).strftime('%d.%m')
            end_date = datetime.fromisoformat(block['end_date']).strftime('%d.%m')
            
            cell = ws.cell(row=6, column=i, value=f"Нед.{week_id}\\n{start_date}-{end_date}")
            cell.font = Font(size=8, bold=True)
            cell.fill = self.info_fill
            cell.alignment = self.center_align
            cell.border = self.border
        
        # Заполняем пакеты работ
        current_row = 7
        for i, package in enumerate(work_packages, 1):
            package_id = package.get('package_id', '')
            
            # Номер п/п
            ws.cell(row=current_row, column=1, value=i).alignment = self.center_align
            
            # Название пакета
            package_name = package.get('name', 'Безымянный пакет')
            ws.cell(row=current_row, column=2, value=package_name).alignment = self.left_align
            
            # Единица измерения и количество - универсальная логика
            volume_data = package.get('volume_data', {})
            calculations = package.get('calculations', {})
            
            # Приоритет: volume_data -> calculations -> дефолт
            unit = (volume_data.get('final_unit') or 
                   volume_data.get('unit') or 
                   calculations.get('final_unit') or 
                   calculations.get('unit') or 'шт')
                   
            quantity = (volume_data.get('final_quantity') or 
                       volume_data.get('quantity') or 
                       calculations.get('final_quantity') or 
                       calculations.get('quantity') or 0)
            
            ws.cell(row=current_row, column=3, value=unit).alignment = self.center_align
            ws.cell(row=current_row, column=4, value=str(quantity)).alignment = self.center_align
            
            # Получаем данные из scheduling_data (приоритет) или из package
            schedule_info = scheduling_data.get(package_id, {})
            schedule_blocks = schedule_info.get('schedule_blocks', package.get('schedule_blocks', []))
            progress_per_block = schedule_info.get('progress_per_block', package.get('progress_per_block', {}))
            staffing_per_block = schedule_info.get('staffing_per_block', package.get('staffing_per_block', {}))
            
            # Даты начала и окончания
            if schedule_blocks:
                start_date = self._get_package_start_date(schedule_blocks, timeline_blocks)
                end_date = self._get_package_end_date(schedule_blocks, timeline_blocks)
                ws.cell(row=current_row, column=5, value=start_date).alignment = self.center_align
                ws.cell(row=current_row, column=6, value=end_date).alignment = self.center_align
            
            # Заполняем прогресс по неделям с улучшенной визуализацией
            for j, block in enumerate(timeline_blocks, 7):
                week_id = block.get('week_id', block.get('block_id'))
                week_str = str(week_id)
                
                if week_id in schedule_blocks and week_str in progress_per_block:
                    progress = progress_per_block[week_str]
                    staffing = staffing_per_block.get(week_str, 0)
                    
                    # Формат: "50%/3чел"
                    cell_value = f"{progress}%/{staffing}чел"
                    cell = ws.cell(row=current_row, column=j, value=cell_value)
                    cell.alignment = self.center_align
                    
                    # Цветовое кодирование по прогрессу
                    if progress >= 70:
                        cell.fill = self.progress_high
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif progress >= 30:
                        cell.fill = self.progress_medium
                        cell.font = Font(color="FFFFFF", bold=True)
                    else:
                        cell.fill = self.progress_low
                        cell.font = Font(color="FFFFFF", bold=True)
                    
                    cell.border = self.border
                else:
                    # Пустая ячейка с границами
                    cell = ws.cell(row=current_row, column=j, value="")
                    cell.border = self.border
            
            current_row += 1
        
        # Применяем форматирование
        self._format_schedule_sheet(ws, timeline_cols)
    
    def _create_reasoning_sheet(self, ws, work_packages: List[Dict], timeline_blocks: List[Dict], project_info: Dict, scheduling_data: Dict):
        """Создает лист с планированием и обоснованиями"""
        
        # Заголовок
        ws['A1'] = "📅 ПЛАНИРОВАНИЕ И ОБОСНОВАНИЯ РЕШЕНИЙ"
        ws.merge_cells('A1:F1')
        ws['A1'].font = Font(bold=True, size=16, color="2E4057")
        ws['A1'].alignment = self.center_align
        ws['A1'].fill = self.reasoning_fill
        
        # Подзаголовок
        ws['A2'] = "Детальные обоснования планирования работ по пакетам"
        ws.merge_cells('A2:F2')
        ws['A2'].font = Font(bold=True, size=12, color="4A628A")
        ws['A2'].alignment = self.center_align
        
        current_row = 4
        
        # Обрабатываем только пакеты, для которых есть scheduling_reasoning
        scheduled_packages = []
        for package in work_packages:
            package_id = package.get('package_id', '')
            if package_id in scheduling_data:
                scheduled_packages.append(package)
        
        if not scheduled_packages:
            ws.cell(row=current_row, column=1, value="⚠️ Обоснования планирования не найдены").font = Font(bold=True, color="F44336")
            return
        
        for i, package in enumerate(scheduled_packages, 1):
            package_id = package.get('package_id', '')
            schedule_info = scheduling_data.get(package_id, {})
            reasoning = schedule_info.get('scheduling_reasoning', {})
            
            if not reasoning:
                continue
            
            # Заголовок пакета
            package_header = f"📦 ПАКЕТ {i}: {package.get('name', 'Безымянный пакет')}"
            ws.cell(row=current_row, column=1, value=package_header)
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=14, color="2E4057")
            ws.cell(row=current_row, column=1).fill = self.subheader_fill
            ws.merge_cells(f'A{current_row}:F{current_row}')
            current_row += 1
            
            # Основная информация
            schedule_blocks = schedule_info.get('schedule_blocks', [])
            progress_per_block = schedule_info.get('progress_per_block', {})
            staffing_per_block = schedule_info.get('staffing_per_block', {})
            
            info_items = [
                ("📅 ID пакета:", package_id),
                ("⏱️ Недели выполнения:", f"{min(schedule_blocks)}-{max(schedule_blocks)}" if schedule_blocks else "Не определено"),
                ("👥 Общая численность:", f"{sum(staffing_per_block.values())} чел·нед" if staffing_per_block else "0"),
                ("📈 Распределение прогресса:", " | ".join([f"Нед.{k}: {v}%" for k, v in progress_per_block.items()]) if progress_per_block else "Не определено")
            ]
            
            for label, value in info_items:
                ws.cell(row=current_row, column=1, value=label).font = Font(bold=True, size=10)
                ws.cell(row=current_row, column=2, value=str(value))
                ws.cell(row=current_row, column=1).fill = self.info_fill
                ws.cell(row=current_row, column=2).fill = self.info_fill
                current_row += 1
            
            current_row += 1
            
            # Обоснования
            reasoning_items = [
                ("🗓️ ПОЧЕМУ ИМЕННО ЭТИ НЕДЕЛИ:", reasoning.get('why_these_weeks', 'Не указано')),
                ("⏳ ПОЧЕМУ ИМЕННО ТАКАЯ ПРОДОЛЖИТЕЛЬНОСТЬ:", reasoning.get('why_this_duration', 'Не указано')),
                ("📊 ПОЧЕМУ ИМЕННО ТАКАЯ ПОСЛЕДОВАТЕЛЬНОСТЬ:", reasoning.get('why_this_sequence', 'Не указано')),
                ("👷 ПОЧЕМУ ИМЕННО ТАКОЕ КОЛИЧЕСТВО ЛЮДЕЙ:", reasoning.get('why_this_staffing', 'Не указано'))
            ]
            
            for j, (label, explanation) in enumerate(reasoning_items):
                # Заголовок обоснования
                ws.cell(row=current_row, column=1, value=label)
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=11, color="FF6F00")
                ws.cell(row=current_row, column=1).fill = self.reasoning_fill
                ws.merge_cells(f'A{current_row}:F{current_row}')
                current_row += 1
                
                # Текст обоснования
                ws.cell(row=current_row, column=1, value=explanation)
                ws.cell(row=current_row, column=1).alignment = self.top_left_align
                ws.cell(row=current_row, column=1).font = Font(size=10)
                ws.merge_cells(f'A{current_row}:F{current_row}')
                
                # Устанавливаем высоту строки для длинного текста
                ws.row_dimensions[current_row].height = max(20, len(explanation) // 100 * 15)
                
                current_row += 2
            
            current_row += 2  # Пропуск между пакетами
        
        # Применяем форматирование
        self._format_reasoning_sheet(ws)
    
    def _create_packages_sheet(self, ws, work_packages: List[Dict], project_info: Dict):
        """Создает лист с детальной информацией по пакетам работ"""
        
        # Заголовок
        ws['A1'] = "📋 ПАКЕТЫ РАБОТ - ДЕТАЛЬНАЯ ИНФОРМАЦИЯ"
        ws.merge_cells('A1:H1')
        ws['A1'].font = Font(bold=True, size=16, color="366092")
        ws['A1'].alignment = self.center_align
        
        current_row = 3
        
        # Создаем детальную информацию для каждого пакета
        for i, package in enumerate(work_packages, 1):
            # Заголовок пакета
            package_header = f"📦 ПАКЕТ {i}: {package.get('name', 'Безымянный пакет')}"
            ws.cell(row=current_row, column=1, value=package_header).font = Font(bold=True, size=14, color="366092")
            ws.merge_cells(f'A{current_row}:H{current_row}')
            current_row += 1
            
            # Основная информация пакета
            volume_data = package.get('volume_data', {})
            calculations = package.get('calculations', {})
            
            unit = (volume_data.get('final_unit') or volume_data.get('unit') or 
                   calculations.get('final_unit') or calculations.get('unit') or 'шт')
            quantity = (volume_data.get('final_quantity') or volume_data.get('quantity') or 
                       calculations.get('final_quantity') or calculations.get('quantity') or 0)
            
            # Детали пакета
            package_details = [
                ("ID пакета:", package.get('package_id', 'N/A')),
                ("Единица измерения:", unit),
                ("Количество:", str(quantity)),
                ("Описание:", package.get('description', 'Не указано'))
            ]
            
            for label, value in package_details:
                ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=value)
                current_row += 1
            
            # Логика расчета и обоснования
            logic = (volume_data.get('calculation_logic') or 
                    calculations.get('calculation_logic') or 
                    'Логика расчета не указана')
            
            ws.cell(row=current_row, column=1, value="Логика расчета:").font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=logic[:200] + "..." if len(logic) > 200 else logic)
            current_row += 1
            
            # Подробные обоснования
            reasoning = volume_data.get('reasoning', {})
            if reasoning:
                current_row += 1
                ws.cell(row=current_row, column=1, value="📝 ОБОСНОВАНИЯ РАСЧЕТОВ:").font = Font(bold=True, size=11, color="2F5233")
                current_row += 1
                
                # Обоснование количества
                why_quantity = reasoning.get('why_this_quantity', '')
                if why_quantity:
                    ws.cell(row=current_row, column=1, value="• Почему это количество:").font = Font(bold=True)
                    ws.merge_cells(f'B{current_row}:H{current_row}')
                    ws.cell(row=current_row, column=2, value=why_quantity).alignment = self.left_align
                    current_row += 1
                
                # Обоснование единицы измерения
                why_unit = reasoning.get('why_this_unit', '')
                if why_unit:
                    ws.cell(row=current_row, column=1, value="• Почему эта единица:").font = Font(bold=True)
                    ws.merge_cells(f'B{current_row}:H{current_row}')
                    ws.cell(row=current_row, column=2, value=why_unit).alignment = self.left_align
                    current_row += 1
                
                # Подход к расчету
                calc_approach = reasoning.get('calculation_approach', '')
                if calc_approach:
                    ws.cell(row=current_row, column=1, value="• Подход к расчету:").font = Font(bold=True)
                    ws.merge_cells(f'B{current_row}:H{current_row}')
                    ws.cell(row=current_row, column=2, value=calc_approach).alignment = self.left_align
                    current_row += 1
            
            # Список работ в пакете
            current_row += 1
            ws.cell(row=current_row, column=1, value="📋 ВХОДЯЩИЕ РАБОТЫ:").font = Font(bold=True, size=12, color="2F5233")
            current_row += 1
            
            # Заголовки для работ
            work_headers = ["№", "Код работы", "Наименование работы", "Единица", "Количество", "Роль", "Участие"]
            for col, header in enumerate(work_headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = self.info_fill
                cell.alignment = self.center_align
                cell.border = self.border
            current_row += 1
            
            # Найдем работы этого пакета
            package_works = self._get_package_works(package.get('package_id'), work_packages, project_info)
            
            for j, work in enumerate(package_works, 1):
                ws.cell(row=current_row, column=1, value=j).alignment = self.center_align
                ws.cell(row=current_row, column=2, value=work.get('code', 'N/A'))
                ws.cell(row=current_row, column=3, value=work.get('name', 'Без названия'))
                ws.cell(row=current_row, column=4, value=work.get('unit', 'шт')).alignment = self.center_align
                ws.cell(row=current_row, column=5, value=str(work.get('quantity', 0))).alignment = self.right_align
                ws.cell(row=current_row, column=6, value=work.get('role', 'основная')).alignment = self.center_align
                ws.cell(row=current_row, column=7, value=work.get('included', 'полная')).alignment = self.center_align
                
                # Применяем границы
                for col in range(1, 8):
                    ws.cell(row=current_row, column=col).border = self.border
                    
                current_row += 1
            
            current_row += 2  # Пропуск между пакетами
        
        # Применяем форматирование
        self._format_packages_sheet(ws)
    
    def _get_package_works(self, package_id, work_packages, project_info):
        """Получает список работ для указанного пакета"""
        works = []
        
        # Сначала ищем в source_work_items
        source_works = project_info.get('source_work_items', [])
        for work in source_works:
            if work.get('package_id') == package_id:
                works.append({
                    'code': work.get('code', 'N/A'),
                    'name': work.get('name', 'Без названия'),
                    'unit': work.get('unit', 'шт'),
                    'quantity': work.get('quantity', 0),
                    'role': 'исходная работа'
                })
        
        # Если не нашли в source_work_items, ищем в volume_data пакета
        if not works:
            for pkg in work_packages:
                if pkg.get('package_id') == package_id:
                    volume_data = pkg.get('volume_data', {})
                    component_analysis = volume_data.get('component_analysis', [])
                    
                    for component in component_analysis:
                        works.append({
                            'code': component.get('code', 'N/A'),
                            'name': component.get('work_name', 'Без названия'),
                            'unit': component.get('unit', 'шт'),
                            'quantity': component.get('quantity', 0),
                            'role': self._translate_role(component.get('role', 'unknown')),
                            'included': self._translate_included(component.get('included', 'full'))
                        })
                    break
        
        return works
    
    def _translate_role(self, role):
        """Переводит роль работы на русский язык"""
        role_translations = {
            'base_surface': 'базовая поверхность',
            'finish_layer': 'финишный слой', 
            'adjustment': 'корректировка',
            'preparation': 'подготовка',
            'base_element': 'базовый элемент',
            'safety_element': 'элемент безопасности',
            'separate_work': 'отдельная работа',
            'full': 'полная',
            'excluded': 'исключена',
            'reference': 'справочная'
        }
        return role_translations.get(role, role)
    
    def _translate_included(self, included):
        """Переводит тип участия работы в расчетах на русский язык"""
        included_translations = {
            'full': 'полная',
            'excluded': 'исключена',
            'reference': 'справочная',
            'partial': 'частичная'
        }
        return included_translations.get(included, included)
    
    def _create_logic_sheet(self, ws, work_packages: List[Dict], project_info: Dict, extracted_data: Dict):
        """Создает лист с логикой расчетов и технической информацией"""
        
        # Заголовок
        ws['A1'] = "🧮 ЛОГИКА РАСЧЕТОВ И ТЕХНИЧЕСКАЯ ИНФОРМАЦИЯ"
        ws.merge_cells('A1:E1')
        ws['A1'].font = Font(bold=True, size=16, color="366092")
        ws['A1'].alignment = self.center_align
        
        current_row = 3
        
        # Блок 1: Общая информация о проекте
        ws.cell(row=current_row, column=1, value="📊 ИНФОРМАЦИЯ О ПРОЕКТЕ").font = Font(bold=True, size=12)
        current_row += 1
        
        project_details = [
            ("Название проекта:", project_info.get('project_name', 'Безымянный проект')),
            ("Структура данных:", project_info.get('structure_version', '1.0')),
            ("Дата создания:", self._format_datetime(project_info.get('created_at'))),
            ("Всего пакетов работ:", len(work_packages)),
            ("Всего временных блоков:", len(extracted_data.get('timeline_blocks', [])))
        ]
        
        for label, value in project_details:
            ws.cell(row=current_row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=current_row, column=2, value=str(value))
            current_row += 1
        
        current_row += 2
        
        # Блок 2: Логика расчетов для каждого пакета
        ws.cell(row=current_row, column=1, value="🧮 ДЕТАЛЬНАЯ ЛОГИКА РАСЧЕТОВ").font = Font(bold=True, size=12)
        current_row += 1
        
        # Заголовки таблицы
        calc_headers = ["Пакет", "Единица", "Количество", "Логика расчета"]
        for col, header in enumerate(calc_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.logic_fill
            cell.alignment = self.center_align
            cell.border = self.border
        current_row += 1
        
        # Заполняем логику расчетов
        for package in work_packages:
            package_name = package.get('name', 'Безымянный пакет')
            
            # Данные расчетов - универсальная логика
            volume_data = package.get('volume_data', {})
            calculations = package.get('calculations', {})
            
            # Приоритет: volume_data -> calculations -> дефолт
            unit = (volume_data.get('final_unit') or 
                   volume_data.get('unit') or 
                   calculations.get('final_unit') or 
                   calculations.get('unit') or 'шт')
                   
            quantity = (volume_data.get('final_quantity') or 
                       volume_data.get('quantity') or 
                       calculations.get('final_quantity') or 
                       calculations.get('quantity') or 0)
            
            # Логика расчета и обоснования
            logic = (volume_data.get('calculation_logic') or 
                    calculations.get('calculation_logic') or 
                    calculations.get('calculation_summary') or 
                    'Логика расчета не указана')
                    
            reasoning = volume_data.get('reasoning', {})
            
            # Расширенная логика с обоснованиями (только если есть volume_data)
            if reasoning and volume_data:
                why_quantity = reasoning.get('why_this_quantity', '')
                why_unit = reasoning.get('why_this_unit', '')
                approach = reasoning.get('calculation_approach', '')
                
                if why_quantity or why_unit or approach:
                    extended_logic = f"{logic}\n\nОбоснование количества: {why_quantity}\nОбоснование единицы: {why_unit}\nПодход к расчету: {approach}"
                    logic = extended_logic[:400] + "..." if len(extended_logic) > 400 else extended_logic
            
            ws.cell(row=current_row, column=1, value=package_name).alignment = self.left_align
            ws.cell(row=current_row, column=2, value=unit).alignment = self.center_align
            ws.cell(row=current_row, column=3, value=str(quantity)).alignment = self.right_align
            ws.cell(row=current_row, column=4, value=logic).alignment = self.left_align
            
            current_row += 1
        
        current_row += 2
        
        # Блок 3: Статус pipeline
        pipeline_status = extracted_data.get('pipeline_status', {})
        if pipeline_status:
            ws.cell(row=current_row, column=1, value="🔄 СТАТУС ОБРАБОТКИ").font = Font(bold=True, size=12)
            current_row += 1
            
            # Для v2.0
            if isinstance(pipeline_status, dict):
                agents = pipeline_status.get('agents_status', [])
                current_stage = pipeline_status.get('current_stage', 'unknown')
                
                ws.cell(row=current_row, column=1, value="Текущая стадия:").font = Font(bold=True)
                ws.cell(row=current_row, column=2, value=current_stage)
                current_row += 1
                
                for agent in agents:
                    agent_name = agent.get('agent', 'unknown')
                    status = agent.get('status', 'unknown')
                    duration = agent.get('duration', 'N/A')
                    
                    ws.cell(row=current_row, column=1, value=f"{agent_name}:")
                    ws.cell(row=current_row, column=2, value=status)
                    ws.cell(row=current_row, column=3, value=f"{duration}s" if duration and duration != 'N/A' else 'N/A')
                    current_row += 1
            
            # Для v1.0
            elif isinstance(pipeline_status, list):
                for status in pipeline_status:
                    agent_name = status.get('agent_name', 'unknown')
                    agent_status = status.get('status', 'unknown')
                    
                    ws.cell(row=current_row, column=1, value=f"{agent_name}:")
                    ws.cell(row=current_row, column=2, value=agent_status)
                    current_row += 1
        
        # Применяем форматирование
        self._format_logic_sheet(ws)
    
    def _get_package_start_date(self, schedule_blocks: List[int], timeline_blocks: List[Dict]) -> str:
        """Получает дату начала пакета работ"""
        if not schedule_blocks:
            return ""
        
        min_week = min(schedule_blocks)
        for block in timeline_blocks:
            block_id = block.get('week_id', block.get('block_id'))
            if block_id == min_week:
                date = datetime.fromisoformat(block['start_date'])
                return date.strftime('%d.%m.%Y')
        return ""
    
    def _get_package_end_date(self, schedule_blocks: List[int], timeline_blocks: List[Dict]) -> str:
        """Получает дату окончания пакета работ"""
        if not schedule_blocks:
            return ""
        
        max_week = max(schedule_blocks)
        for block in timeline_blocks:
            block_id = block.get('week_id', block.get('block_id'))
            if block_id == max_week:
                date = datetime.fromisoformat(block['end_date'])
                return date.strftime('%d.%m.%Y')
        return ""
    
    def _format_datetime(self, datetime_str: str) -> str:
        """Форматирует datetime в читаемый вид"""
        if not datetime_str:
            return "Не указано"
        try:
            dt = datetime.fromisoformat(datetime_str.replace('Z', '+00:00'))
            return dt.strftime('%d.%m.%Y %H:%M')
        except:
            return str(datetime_str)
    
    def _format_schedule_sheet(self, ws, timeline_cols: int):
        """Применяет форматирование к листу календарного графика"""
        
        # Ширина колонок
        ws.column_dimensions['A'].width = 5   # №
        ws.column_dimensions['B'].width = 40  # Наименование работ
        ws.column_dimensions['C'].width = 8   # Ед.изм
        ws.column_dimensions['D'].width = 10  # Количество
        ws.column_dimensions['E'].width = 12  # Начало
        ws.column_dimensions['F'].width = 12  # Окончание
        
        # Временные колонки
        for i in range(7, 7 + timeline_cols):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 10
        
        # Применяем границы ко всем ячейкам с данными
        max_row = ws.max_row
        max_col = 6 + timeline_cols
        
        for row in range(5, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.border
    
    def _format_packages_sheet(self, ws):
        """Применяет форматирование к листу пакетов работ"""
        
        # Ширина колонок
        ws.column_dimensions['A'].width = 5   # №
        ws.column_dimensions['B'].width = 20  # Код работы
        ws.column_dimensions['C'].width = 50  # Наименование работы
        ws.column_dimensions['D'].width = 12  # Единица
        ws.column_dimensions['E'].width = 12  # Количество
        ws.column_dimensions['F'].width = 18  # Роль
        ws.column_dimensions['G'].width = 12  # Участие
        ws.column_dimensions['H'].width = 15  # Дополнительная колонка
        
        # Границы для всех ячеек
        max_row = ws.max_row
        for row in range(3, max_row + 1):
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = self.border
    
    def _format_reasoning_sheet(self, ws):
        """Применяет форматирование к листу планирования и обоснований"""
        
        # Ширина колонок
        ws.column_dimensions['A'].width = 80  # Основной текст
        ws.column_dimensions['B'].width = 25  # Значения
        ws.column_dimensions['C'].width = 15  # Доп. колонки
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # Применяем границы к заполненным ячейкам
        max_row = ws.max_row
        for row in range(1, max_row + 1):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell.border = self.border
    
    def _format_logic_sheet(self, ws):
        """Применяет форматирование к листу логики расчетов"""
        
        # Ширина колонок
        ws.column_dimensions['A'].width = 30  # Пакет
        ws.column_dimensions['B'].width = 12  # Единица
        ws.column_dimensions['C'].width = 10  # Количество
        ws.column_dimensions['D'].width = 80  # Логика
        ws.column_dimensions['E'].width = 15  # Доп. информация


# Обновленная функция для использования в пайплайне
def generate_multipage_excel_report(input_file: str, output_path: str) -> str:
    """
    Генерация многостраничного Excel отчета в новом формате
    
    Args:
        input_file: Путь к файлу true.json
        output_path: Папка для сохранения
        
    Returns:
        Путь к созданному файлу
    """
    generator = MultiPageScheduleGenerator()
    return generator.generate_multipage_excel(input_file, output_path)


if __name__ == "__main__":
    # Тестирование на реальных данных
    test_input = "/home/imort/Herzog_v3/projects/34975055/d490876a/true.json"
    test_output = "/tmp"
    
    if os.path.exists(test_input):
        print("🧪 Тестирование нового многостраничного отчета с scheduling_reasoning...")
        try:
            result_file = generate_multipage_excel_report(test_input, test_output)
            print(f"✅ Многостраничный отчет создан: {result_file}")
        except Exception as e:
            print(f"❌ Ошибка: {e}")
            import traceback
            traceback.print_exc()
    else:
        print(f"❌ Тестовый файл не найден: {test_input}")